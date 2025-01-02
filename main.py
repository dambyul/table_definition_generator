import pandas as pd
from sqlalchemy import create_engine
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

current_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def read_settings_from_excel(settings_file):
    wb = load_workbook(settings_file, data_only=True)
    ws = wb["Settings"]

    db_config = {
        "system": ws["F12"].value,
        "host": ws["F13"].value,
        "port": ws["F14"].value,
        "user": ws["F15"].value,
        "password": ws["F16"].value,
        "schemas": [s.strip() for s in ws["F17"].value.split(",")],
        "main_color": ws["F21"].value,
        "sub_color" : ws["F22"].value,
    }

    return db_config

def create_excel_with_format(db_config, output_file, main_color, sub_color):
    engine = create_engine(
        f"mysql+pymysql://{db_config['user']}:{db_config['password']}@{db_config['host']}:{db_config['port']}"
    )

    wb = Workbook()
    default_sheet = wb.active
    default_sheet.title = "Default"

    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    thick_border = Border(
        left=Side(style="thick"),
        right=Side(style="thick"),
        top=Side(style="thick"),
        bottom=Side(style="thick"),
    )

    header_fill = PatternFill(start_color=main_color, end_color=main_color, fill_type="solid")
    subheader_fill = PatternFill(start_color=sub_color, end_color=sub_color, fill_type="solid")

    for schema_name in db_config["schemas"]:
        ws = wb.create_sheet(title=schema_name)

        query = """
            SELECT t.TABLE_SCHEMA AS `Database`,
                   t.TABLE_NAME AS `Table`,
                   t.TABLE_COMMENT AS `Table Comment`,
                   c.COLUMN_NAME AS `Column`,
                   c.ORDINAL_POSITION AS `No`,
                   c.COLUMN_TYPE AS `Type & Length`,
                   c.IS_NULLABLE AS `Not Null`,
                   c.COLUMN_KEY AS `Key Type`,
                   c.COLUMN_COMMENT AS `Comment`
            FROM INFORMATION_SCHEMA.TABLES t
                     JOIN
                 INFORMATION_SCHEMA.COLUMNS c
                 ON
                     t.TABLE_SCHEMA = c.TABLE_SCHEMA AND t.TABLE_NAME = c.TABLE_NAME
            WHERE t.TABLE_SCHEMA = %s
            ORDER BY t.TABLE_NAME, c.ORDINAL_POSITION;
        """
        df = pd.read_sql(query, engine, params=(schema_name,))

        row_num = 1
        for table_name, group in df.groupby("Table"):
            first_row = row_num

            ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=4)
            ws.cell(row=row_num, column=1).value = "스키마 명"
            ws.cell(row=row_num, column=1).alignment = header_alignment
            ws.cell(row=row_num, column=1).font = header_font
            ws.cell(row=row_num, column=1).fill = header_fill
            ws.cell(row=row_num, column=2).value = schema_name

            ws.merge_cells(start_row=row_num, start_column=6, end_row=row_num, end_column=8)
            ws.cell(row=row_num, column=5).value = "테이블 명"
            ws.cell(row=row_num, column=5).alignment = header_alignment
            ws.cell(row=row_num, column=5).font = header_font
            ws.cell(row=row_num, column=5).fill = header_fill
            ws.cell(row=row_num, column=6).value = table_name

            ws.merge_cells(start_row=row_num + 1, start_column=2, end_row=row_num + 1, end_column=4)
            ws.cell(row=row_num + 1, column=1).value = "작성일자"
            ws.cell(row=row_num + 1, column=1).alignment = header_alignment
            ws.cell(row=row_num + 1, column=1).font = header_font
            ws.cell(row=row_num + 1, column=1).fill = header_fill
            ws.cell(row=row_num + 1, column=2).value = current_timestamp

            ws.merge_cells(start_row=row_num + 1, start_column=6, end_row=row_num + 1, end_column=8)
            ws.cell(row=row_num + 1, column=5).value = "테이블 설명"
            ws.cell(row=row_num + 1, column=5).alignment = header_alignment
            ws.cell(row=row_num + 1, column=5).font = header_font
            ws.cell(row=row_num + 1, column=5).fill = header_fill
            ws.cell(row=row_num + 1, column=6).value = group["Table Comment"].iloc[0]

            for col in range(1, 9):
                for row in range(row_num, row_num + 2):
                    ws.cell(row=row, column=col).border = border_style

                current_border = ws.cell(row=row_num + 1, column=col).border
                ws.cell(row=row_num + 1, column=col).border = Border(
                    left=current_border.left,
                    right=current_border.right,
                    top=current_border.top,
                    bottom=thick_border.bottom
                )

            row_num += 3

            column_headers = [
                "컬럼명", "No", "컬럼 ID", "타입 및 길이", "Not Null", "PK", "IDX", "비고"
            ]
            for col_num, header in enumerate(column_headers, start=1):
                cell = ws.cell(row=row_num, column=col_num, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
                cell.fill = subheader_fill
                cell.border = Border(
                    top=thick_border.top,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=thick_border.bottom,
                )

            row_num += 1

            is_first_row = True

            for _, row in group.iterrows():
                top_border = thick_border.top if is_first_row else border_style.top
                ws.cell(row=row_num, column=1, value=row["Comment"]).border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=2, value=row["No"]).border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=3, value=row["Column"]).border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=4, value=row["Type & Length"]).border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=5, value="YES" if row["Not Null"] == "NO" else "NO").border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=6, value="PRI" if row["Key Type"] == "PRI" else None).border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=7, value="MUL" if row["Key Type"] == "MUL" else None).border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                ws.cell(row=row_num, column=8, value="").border = Border(
                    top=top_border,
                    left=border_style.left,
                    right=border_style.right,
                    bottom=border_style.bottom,
                )
                if is_first_row:
                    is_first_row = False

                row_num += 1

            for i in range(first_row, row_num):
                for j in range(1, 9): 
                    current_border = ws.cell(row=i, column=j).border
                    ws.cell(row=i, column=j).border = Border(
                        left=thick_border.left if j == 1 else current_border.left,
                        right=thick_border.right if j == 8 else current_border.right,
                        top=thick_border.top if i == first_row else current_border.top,
                        bottom=thick_border.bottom if i == row_num - 1 else current_border.bottom
                    )

            row_num += 1

        for col_index, col_cells in enumerate(ws.iter_cols(), start=1):
            col_letter = get_column_letter(col_index)

            if col_letter == "H" or col_letter == 'E':
                ws.column_dimensions[col_letter].width = 15
                continue

            max_length = 0
            for cell in col_cells:
                if cell.coordinate in ws.merged_cells: 
                    continue

                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(output_file)

settings_file = "settings.xlsx"
#settings_file = "test/settings.xlsx"

output_file = "table_definitions.xlsx"

db_config = read_settings_from_excel(settings_file)
create_excel_with_format(db_config, output_file, db_config["main_color"], db_config["sub_color"])

print("출력 완료료")